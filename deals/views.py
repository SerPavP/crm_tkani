from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from .models import Deal, DealItem
from .forms import DealForm, DealItemForm
from clients.models import Client
from fabrics.models import FabricColor, Fabric
from core.models import ActivityLog
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment
from decimal import Decimal
import datetime


def format_price_for_display(value):
    """
    Форматирует цену с разделителями тысяч для отображения
    Пример: 183227.49 -> 183 227,49
    """
    if value is None:
        return "0,00"
    
    try:
        value = float(value)
        # Форматируем с двумя знаками после запятой
        formatted = f"{value:.2f}"
        
        # Разделяем на целую и дробную части
        if '.' in formatted:
            integer_part, decimal_part = formatted.split('.')
        else:
            integer_part = formatted
            decimal_part = '00'
        
        # Добавляем разделители тысяч в целую часть
        if len(integer_part) > 3:
            # Разбиваем на группы по 3 цифры справа налево
            groups = []
            for i in range(len(integer_part), 0, -3):
                start = max(0, i - 3)
                groups.insert(0, integer_part[start:i])
            integer_part = ' '.join(groups)
        
        # Возвращаем результат с запятой в качестве десятичного разделителя
        return f"{integer_part},{decimal_part}"
    
    except (ValueError, TypeError):
        return "0,00"


@login_required
def deal_list(request):
    """Список всех сделок"""
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    
    deals = Deal.objects.all().order_by('-created_at')
    
    if search_query:
        deals = deals.filter(
            Q(deal_number__icontains=search_query) |
            Q(client__nickname__icontains=search_query) |
            Q(client__phone__icontains=search_query)
        )
    
    if status_filter:
        deals = deals.filter(status=status_filter)
    
    context = {
        'deals': deals,
        'search_query': search_query,
        'status_filter': status_filter,
        'status_choices': Deal.STATUS_CHOICES,
    }
    return render(request, 'deals/deal_list.html', context)


@login_required
def deal_detail(request, deal_id):
    """Детальная информация о сделке"""
    deal = get_object_or_404(Deal, id=deal_id)
    
    # Обновляем общие суммы сделки, если они не установлены
    if deal.total_amount is None:
        deal.update_totals()
        deal.save()
    
    deal_items = deal.dealitem_set.all()
    fabrics = Fabric.objects.all().order_by('name')
    
    # Рассчитываем прибыль компании
    total_profit = 0
    for item in deal_items:
        if item.fabric_color and item.fabric_color.fabric.cost_price:
            # Прибыль = (Цена продажи - Себестоимость) * Количество метров
            profit_per_meter = item.price_per_meter - item.fabric_color.fabric.cost_price
            item_profit = profit_per_meter * item.width_meters
            total_profit += item_profit
    
    # Скрываем прибыль для бухгалтеров
    if request.user.userprofile.role == 'accountant':
        total_profit = None

    context = {
        'deal': deal,
        'deal_items': deal_items,
        'fabrics': fabrics,
        'total_profit': total_profit,
    }
    return render(request, 'deals/deal_detail.html', context)


@login_required
def deal_create(request):
    """Создание новой сделки"""
    if request.method == 'POST':
        form = DealForm(request.POST)
        if form.is_valid():
            deal = form.save()
            
            # Логирование
            ActivityLog.objects.create(
                user=request.user,
                action='create',
                object_type='Deal',
                object_id=deal.id,
                description=f'Создана сделка: {deal.deal_number}'
            )
            
            messages.success(request, f'Сделка "{deal.deal_number}" успешно создана.')
            return redirect('deals:deal_detail', deal_id=deal.id)
    else:
        form = DealForm()
        
        # Предзаполнение клиента если передан в GET параметрах
        client_id = request.GET.get('client')
        if client_id:
            try:
                client = Client.objects.get(id=client_id)
                form.fields['client'].initial = client
            except Client.DoesNotExist:
                pass
    
    # Поиск клиентов
    search_query = request.GET.get('client_search', '')
    clients = Client.objects.all()
    if search_query:
        clients = clients.filter(
            Q(nickname__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    context = {
        'form': form,
        'title': 'Создать сделку',
        'clients': clients[:10],  # Ограничиваем до 10 результатов
        'client_search': search_query,
        'fabric_colors': FabricColor.objects.select_related('fabric').all(),
    }
    return render(request, 'deals/deal_form.html', context)


@login_required
def deal_edit(request, deal_id):
    """Редактирование сделки"""
    deal = get_object_or_404(Deal, id=deal_id)
    
    if request.method == 'POST':
        form = DealForm(request.POST, instance=deal)
        if form.is_valid():
            deal = form.save()
            
            # Логирование
            ActivityLog.objects.create(
                user=request.user,
                action='update',
                object_type='Deal',
                object_id=deal.id,
                description=f'Отредактирована сделка: {deal.deal_number}'
            )
            
            messages.success(request, f'Сделка "{deal.deal_number}" успешно обновлена.')
            return redirect('deals:deal_detail', deal_id=deal.id)
    else:
        form = DealForm(instance=deal)
    
    context = {
        'form': form,
        'title': 'Редактировать сделку',
        'deal': deal,
        'clients': Client.objects.all(),
        'fabric_colors': FabricColor.objects.select_related('fabric').all(),
    }
    return render(request, 'deals/deal_form.html', context)


@login_required
@require_POST
def deal_delete(request, deal_id):
    """Удаление сделки (только для админа)"""
    if request.user.userprofile.role != 'admin':
        return JsonResponse({'error': 'Нет прав доступа'}, status=403)
    
    deal = get_object_or_404(Deal, id=deal_id)
    deal_number = deal.deal_number
    
    # Логирование
    ActivityLog.objects.create(
        user=request.user,
        action='delete',
        object_type='Deal',
        object_id=deal.id,
        description=f'Удалена сделка: {deal_number}'
    )
    
    # Удаляем сделку
    deal.delete()
    
    return JsonResponse({'success': True, 'message': f'Сделка {deal_number} успешно удалена'})


@login_required
def deal_change_status(request, deal_id):
    """Изменение статуса сделки"""
    if request.method == 'POST':
        deal = get_object_or_404(Deal, id=deal_id)
        new_status = request.POST.get('status')
        
        if new_status in [choice[0] for choice in Deal.STATUS_CHOICES]:
            old_status = deal.get_status_display()
            deal.status = new_status
            deal.save()
            
            # Логирование
            ActivityLog.objects.create(
                user=request.user,
                action='update',
                object_type='Deal',
                object_id=deal.id,
                description=f'Изменен статус сделки {deal.deal_number}: {old_status} → {deal.get_status_display()}'
            )
            
            # Если это AJAX запрос, возвращаем JSON
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': f'Статус сделки изменен на "{deal.get_status_display()}".',
                    'new_status': new_status,
                    'new_status_display': deal.get_status_display()
                })
            else:
                messages.success(request, f'Статус сделки изменен на "{deal.get_status_display()}".')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': 'Неверный статус.'
                }, status=400)
            else:
                messages.error(request, 'Неверный статус.')
    
    # Если это AJAX запрос, возвращаем JSON
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'success': False,
            'error': 'Метод не поддерживается.'
        }, status=405)
    
    return redirect('deals:deal_detail', deal_id=deal_id)


@login_required
def add_deal_item(request, deal_id):
    """Добавление новой позиции в сделку"""
    deal = get_object_or_404(Deal, id=deal_id)
    
    if request.method == 'POST':
        form = DealItemForm(request.POST)
        if form.is_valid():
            deal_item = form.save(commit=False)
            deal_item.deal = deal

            # Получаем себестоимость для валидации
            fabric_color = deal_item.fabric_color
            cost_price = fabric_color.fabric.cost_price

            # Валидация для бухгалтера
            if request.user.userprofile.role == 'accountant':
                # Получаем текущую цену ткани (продажная цена или себестоимость)
                current_price = fabric_color.fabric.selling_price or fabric_color.fabric.cost_price
                # Минимальная цена = текущая цена - 30%
                min_price = current_price * Decimal('0.70')
                if deal_item.price_per_meter < min_price:
                    messages.error(request, f'Цена за метр ({deal_item.price_per_meter:.2f} ₸) не может быть ниже текущей цены - 30% ({min_price:.2f} ₸). Поднимите цену!')
                    # Если это AJAX запрос, возвращаем JSON ответ с ошибкой
                    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'error': f'Цена за метр ({deal_item.price_per_meter:.2f} ₸) не может быть ниже текущей цены - 30% ({min_price:.2f} ₸). Поднимите цену!'}, status=400)
                    # Если это обычный запрос, возвращаемся к форме
                    return redirect('deals:deal_detail', deal_id=deal.id) # Или render с формой и ошибками

            deal_item.save()
            
            # Обновляем общие суммы сделки
            deal.update_totals()
            deal.save()
            
            # Логирование
            ActivityLog.objects.create(
                user=request.user,
                action='create',
                object_type='DealItem',
                object_id=deal_item.id,
                description=f'Добавлена позиция в сделку {deal.deal_number}: {deal_item.fabric_color.fabric.name} ({deal_item.fabric_color.color_name})'
            )
            
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                # Для AJAX запросов возвращаем JSON успешного ответа
                return JsonResponse({'success': True, 'deal_total': str(deal.total_amount)})
            else:
                messages.success(request, 'Позиция успешно добавлена.')
                return redirect('deals:deal_detail', deal_id=deal.id)
        else:
            # Если форма невалидна и это AJAX запрос
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                errors = form.errors.as_json()
                return JsonResponse({'success': False, 'errors': errors}, status=400)

    # Этот блок будет выполняться только для GET запросов или если форма невалидна (без AJAX)
    # В случае AJAX ошибок, мы уже вернули JsonResponse выше
    form = DealItemForm()
    context = {
        'deal': deal,
        'form': form,
        'fabrics': Fabric.objects.all().order_by('name'), # Для модального окна
    }
    return render(request, 'deals/deal_detail.html', context) # Рендерим обратно детальную страницу сделки


@login_required
def edit_deal_item(request, deal_id, item_id):
    """Редактирование позиции в сделке"""
    deal = get_object_or_404(Deal, id=deal_id)
    deal_item = get_object_or_404(DealItem, id=item_id, deal=deal)
    
    if request.method == 'POST':
        # Собираем данные из тела запроса, так как это AJAX запрос (application/x-www-form-urlencoded)
        fabric_color_id = request.POST.get('fabric_color')
        width_meters = request.POST.get('width_meters')
        price_per_meter = request.POST.get('price_per_meter')

        try:
            fabric_color = FabricColor.objects.get(id=fabric_color_id)
            deal_item.fabric_color = fabric_color
            deal_item.width_meters = float(width_meters)
            deal_item.price_per_meter = float(price_per_meter)
            
            # Получаем себестоимость для валидации
            cost_price = fabric_color.fabric.cost_price

            # Валидация для бухгалтера
            if request.user.userprofile.role == 'accountant':
                # Получаем текущую цену ткани (продажная цена или себестоимость)
                current_price = fabric_color.fabric.selling_price or fabric_color.fabric.cost_price
                # Минимальная цена = текущая цена - 30%
                min_price = current_price * Decimal('0.70')
                if deal_item.price_per_meter < min_price:
                    return JsonResponse({'success': False, 'error': f'Цена за метр ({deal_item.price_per_meter:.2f} ₸) не может быть ниже текущей цены - 30% ({min_price:.2f} ₸). Поднимите цену!'}, status=400)

            deal_item.save()
            
            # Обновляем общие суммы сделки
            deal.update_totals()
            deal.save()
            
            # Логирование
            ActivityLog.objects.create(
                user=request.user,
                action='update',
                object_type='DealItem',
                object_id=deal_item.id,
                description=f'Отредактирована позиция в сделке {deal.deal_number}: {deal_item.fabric_color.fabric.name} ({deal_item.fabric_color.color_name})'
            )
            
            return JsonResponse({'success': True, 'deal_total': str(deal.total_amount)})
        except FabricColor.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Выбранный цвет ткани не найден.'}, status=400)
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Некорректные данные для количества или цены.'}, status=400)
    
    form = DealItemForm(instance=deal_item)
    context = {
        'deal': deal,
        'deal_item': deal_item,
        'form': form,
        'fabrics': Fabric.objects.all().order_by('name'),
    }
    return render(request, 'deals/edit_deal_item.html', context)


@login_required
@require_POST
def delete_deal_item(request, deal_id, item_id):
    """Удаление позиции сделки"""
    deal = get_object_or_404(Deal, id=deal_id)
    item = get_object_or_404(DealItem, id=item_id, deal=deal)
    
    # Логирование
    ActivityLog.objects.create(
        user=request.user,
        action='delete',
        object_type='DealItem',
        object_id=item.id,
        description=f'Удалена позиция из сделки: {deal.deal_number}'
    )
    
    item.delete()
    
    # Пересчитываем общую сумму сделки после удаления
    deal.update_totals()
    deal.save()
    
    messages.success(request, 'Позиция успешно удалена из сделки.')
    return redirect('deals:deal_detail', deal_id=deal.id)


@login_required
def get_fabric_colors(request):
    """AJAX endpoint для получения цветов ткани"""
    fabric_id = request.GET.get('fabric_id')
    
    if not fabric_id:
        return JsonResponse({'colors': []})
    
    try:
        colors = FabricColor.objects.filter(fabric_id=fabric_id).order_by('color_name')
        
        colors_data = []
        for color in colors:
            price = color.fabric.selling_price or color.fabric.cost_price or 0
            colors_data.append({
                'id': color.id,
                'name': color.color_name,
                'number': color.color_number,
                'hex': color.color_hex,
                'price_per_meter': float(price),
                'cost_price': float(color.fabric.cost_price or 0),
                'active_rolls_count': color.active_rolls_count, # Добавляем количество рулонов
            })
        
        return JsonResponse({'colors': colors_data})
    except Exception as e:
        print(f"Ошибка в get_fabric_colors: {e}")
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def get_fabric_color_price(request):
    """AJAX endpoint для получения цены цвета ткани"""
    fabric_color_id = request.GET.get('fabric_color_id')
    
    try:
        fabric_color = FabricColor.objects.get(id=fabric_color_id)
        return JsonResponse({
            'price_per_meter': float(fabric_color.fabric.selling_price or fabric_color.fabric.cost_price or 0),
            'fabric_name': fabric_color.fabric.name,
            'color_name': fabric_color.color_name,
        })
    except FabricColor.DoesNotExist:
        return JsonResponse({'error': 'Цвет ткани не найден'}, status=404)


@login_required
def get_fabric_color_details(request):
    """Возвращает детали выбранного цвета ткани (себестоимость, количество рулонов)"""
    color_id = request.GET.get('color_id')
    if not color_id:
        return JsonResponse({'error': 'Color ID is required'}, status=400)

    try:
        fabric_color = FabricColor.objects.get(id=color_id)
        data = {
            'cost_price': str(fabric_color.fabric.cost_price), # Convert Decimal to string
            'active_rolls_count': fabric_color.active_rolls_count,
        }
        return JsonResponse(data)
    except FabricColor.DoesNotExist:
        return JsonResponse({'error': 'Fabric color not found'}, status=404)


@login_required
def export_deal_pdf(request, deal_id):
    deal = get_object_or_404(Deal, id=deal_id)
    deal_items = deal.dealitem_set.all()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f"inline; filename=\"deal_{deal.deal_number}.pdf\""

    # Используем A4 размер для лучшего отображения
    from reportlab.lib.pagesizes import A4
    doc = SimpleDocTemplate(response, pagesize=A4, leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    
    # Устанавливаем шрифт по умолчанию для всего документа
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.lib.units import cm
    
    # Настройка шрифта - всегда используем Times New Roman
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    # Регистрируем шрифт Times New Roman
    try:
        pdfmetrics.registerFont(TTFont('TimesNewRoman', 'times.ttf'))
        font_name = 'TimesNewRoman'
        bold_font_name = 'TimesNewRoman'
    except:
        try:
            # Пробуем зарегистрировать встроенный шрифт с поддержкой кириллицы
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
            pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))
            font_name = 'STSong-Light'
            bold_font_name = 'STSong-Light'
        except:
            # Если ничего не работает, используем Times
            font_name = 'Times'
            bold_font_name = 'Times'
    
    # Создаем стили с поддержкой кириллицы и автоматическим переносом
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['h1'],
        fontName=font_name,
        fontSize=20,
        alignment=TA_CENTER,
        spaceAfter=20,
        spaceBefore=10,
        leading=24
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=14,
        alignment=TA_LEFT,
        spaceAfter=6,
        leading=16,
        wordWrap='CJK'  # Автоматический перенос для кириллицы
    )
    
    bold_style = ParagraphStyle(
        'CustomBold',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=14,
        alignment=TA_LEFT,
        spaceAfter=6,
        leading=16,
        wordWrap='CJK'
    )
    
    story = []

    # Заголовок документа
    story.append(Paragraph(f"Сделка №{deal.deal_number}", title_style))
    story.append(Spacer(1, 0.3 * inch))

    # Информация о сделке и клиенте
    story.append(Paragraph(f"<b>Клиент:</b> {deal.client.nickname}", normal_style))
    story.append(Paragraph(f"<b>Телефон:</b> {deal.client.phone}", normal_style))
    story.append(Paragraph(f"<b>Статус:</b> {deal.get_status_display()}", normal_style))
    story.append(Paragraph(f"<b>Дата создания:</b> {deal.created_at.strftime('%d.%m.%Y %H:%M')}", normal_style))
    story.append(Spacer(1, 0.3 * inch))

    # Создаем данные для таблицы с поддержкой переноса текста
    table_data = []
    
    # Заголовки таблицы
    headers = ["№", "Ткань", "Цвет", "Количество (м)", "Цена за м", "Сумма"]
    table_data.append(headers)
    
    # Данные позиций
    for i, item in enumerate(deal_items, 1):
        # Создаем параграфы для текста с автоматическим переносом
        # Убираем ограничение длины - текст будет переноситься автоматически
        fabric_para = Paragraph(item.fabric_color.fabric.name, normal_style)
        color_para = Paragraph(f"{item.fabric_color.color_name} (№{item.fabric_color.color_number})", normal_style)
        
        row = [
            str(i),
            fabric_para,
            color_para,
            str(item.width_meters),
            f"{format_price_for_display(item.price_per_meter)} ₸",
            f"{format_price_for_display(item.total_price)} ₸"
        ]
        table_data.append(row)
    
    # Итоговая строка
    table_data.append(["", "", "", "", "Итого:", f"{format_price_for_display(deal.total_amount)} ₸"])

    # Создаем таблицу с фиксированными ширинами колонок - увеличиваем ширину для лучшего отображения
    col_widths = [1*cm, 5*cm, 4*cm, 2.5*cm, 3*cm, 3*cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1, splitByRow=True)
    
    # Улучшенные стили для таблицы
    style = TableStyle([
        # Заголовок таблицы - темно-синий фон с белым текстом
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), bold_font_name),
        ("FONTSIZE", (0, 0), (-1, 0), 14),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
        ("FONTWEIGHT", (0, 0), (-1, 0), "BOLD"),
        
        # Данные таблицы - белый фон с черным текстом
        ("BACKGROUND", (0, 1), (-1, -3), colors.white),
        ("TEXTCOLOR", (0, 1), (-1, -3), colors.black),
        ("FONTNAME", (0, 1), (-1, -1), font_name),
        ("FONTSIZE", (0, 1), (-1, -1), 14),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),  # Номер по центру
        ("ALIGN", (1, 1), (2, -1), "LEFT"),    # Ткань и цвет по левому краю
        ("ALIGN", (3, 1), (5, -1), "CENTER"),  # Количество, цена и сумма по центру
        
        # Итоговая строка - желтый фон с черным текстом
        ("BACKGROUND", (0, -1), (-1, -1), colors.yellow),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.black),
        ("FONTWEIGHT", (4, -1), (5, -1), "BOLD"),
        ("FONTSIZE", (0, -1), (-1, -1), 14),
        
        # Границы
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("LINEBELOW", (0, 0), (-1, 0), 2, colors.black),  # Толстая линия под заголовком
        
        # Отступы
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (1, 1), (-1, -1), 10),
        ("BOTTOMPADDING", (1, 1), (-1, -1), 10),
        
        # Высота строк - увеличиваем для лучшего отображения
        ("MINIMUMHEIGHT", (0, 0), (-1, -1), 25),
    ])
    table.setStyle(style)
    
    story.append(table)
    story.append(Spacer(1, 0.5 * inch))

    doc.build(story)
    return response


@login_required
def export_deal_excel(request, deal_id):
    deal = get_object_or_404(Deal, id=deal_id)
    deal_items = deal.dealitem_set.all()

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=\"deal_{deal.deal_number}.xlsx\""

    wb = Workbook()
    ws = wb.active
    ws.title = f"Сделка {deal.deal_number}"

    # Headers
    headers = ["№", "Ткань", "Цвет", "Количество (м)", "Цена за м", "Сумма"]
    ws.append(headers)
    
    # Apply bold font to headers
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Deal Items
    for i, item in enumerate(deal_items, 1):
        ws.append([
            i,
            item.fabric_color.fabric.name,
            f"{item.fabric_color.color_name} (№{item.fabric_color.color_number})",
            item.width_meters,
            float(item.price_per_meter),
            float(item.total_price)
        ])
        for cell in ws[ws.max_row]:
            cell.font = Font(size=12)
            cell.border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Totals
    ws.append(["", "", "", "", "Итого:", f"{deal.total_amount} ₸"])

    # Apply bold font and borders to total row
    for row_idx in range(ws.max_row, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).font = Font(bold=True, size=14)
            ws.cell(row=row_idx, column=col_idx).border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    # Merge cells for total
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.5
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response


@login_required
def print_deal_warehouse(request, deal_id):
    deal = get_object_or_404(Deal, id=deal_id)
    deal_items = deal.dealitem_set.all()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f"inline; filename=\"warehouse_{deal.deal_number}.pdf\""

    doc = SimpleDocTemplate(response, pagesize=letter)
    styles = getSampleStyleSheet()
    
    # Устанавливаем шрифт по умолчанию для всего документа
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    # Настройка шрифта - всегда используем Times New Roman
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    
    # Всегда используем Times New Roman
    try:
        pdfmetrics.registerFont(TTFont('TimesNewRoman', 'times.ttf'))
        font_name = 'TimesNewRoman'
        bold_font_name = 'TimesNewRoman'
    except:
        # Если Times New Roman не найден, используем Times
        font_name = 'Times'
        bold_font_name = 'Times'
    
    # Создаем стили с поддержкой кириллицы
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['h1'],
        fontName=font_name,
        fontSize=22,
        alignment=TA_CENTER,
        spaceAfter=25
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=14,
        alignment=TA_LEFT,
        spaceAfter=8
    )
    
    story = []

    # Заголовок документа
    story.append(Paragraph(f"Сделка №{deal.deal_number}", title_style))
    story.append(Spacer(1, 0.2 * inch))

    # Таблица позиций (только для склада - без цен)
    data = [["№", "Ткань", "Цвет", "Количество (м)"]]
    for i, item in enumerate(deal_items, 1):
        data.append([
            str(i),
            item.fabric_color.fabric.name,
            f"{item.fabric_color.color_name} (№{item.fabric_color.color_number})",
            str(item.width_meters)
        ])

    table = Table(data)
    
    # Стили для таблицы
    style = TableStyle([
        # Заголовок таблицы - серый фон
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), bold_font_name),
        ("FONTSIZE", (0, 0), (-1, 0), 14),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 16),
        ("TOPPADDING", (0, 0), (-1, 0), 12),
        
        # Данные таблицы - желтый фон
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("FONTNAME", (0, 1), (-1, -1), font_name),
        ("FONTSIZE", (0, 1), (-1, -1), 12),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),  # Номер по центру
        ("ALIGN", (1, 1), (2, -1), "LEFT"),    # Ткань и цвет по левому краю
        ("ALIGN", (3, 1), (3, -1), "CENTER"),  # Количество по центру
        
        # Границы
        ("GRID", (0, 0), (-1, -1), 1.5, colors.black),
        
        # Отступы
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (1, 1), (-1, -1), 12),
        ("BOTTOMPADDING", (1, 1), (-1, -1), 12),
    ])
    table.setStyle(style)
    
    story.append(table)

    doc.build(story)
    return response

